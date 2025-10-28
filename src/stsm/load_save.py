import tkinter as tk
import cv2
import tkinter.filedialog as fd
import tkinter.simpledialog as sd
import tkinter.messagebox as messagebox
import os
from layer_controls import show_layer_controls, hide_layer_controls, hide_proc_layer_controls
from display import update_display, update_proc_display
from roi import start_roi, update_roi, end_roi_drag, confirm_roi, process_selected_roi, set_confirm_roi_button_visible
import openpyxl as opxl
import h5py


def load_image(self):
    # Clear previous images
    self.images = []
    self.tk_images = []
    
    # Hide controls if they were previously shown
    hide_layer_controls(self)
    
    # Open file dialog to select images
    file_paths = fd.askopenfilenames(
        title="Select Two Images",
        filetypes=[("Image files", "*.jpg *.jpeg *.png *.tif *.tiff")]
    )
    
    if len(file_paths) < 2:
        messagebox.showwarning("Warning", "Please select at least two images.")
        return
        
    # Load the first two selected images using OpenCV
    for i, file in enumerate(file_paths[:2]):
        image = cv2.imread(file)
        if image is not None:
            # Convert to grayscale
            gray_image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
            self.images.append(gray_image)
            self.layer_names[i] = os.path.basename(file)  # Use the file name as the layer name
        else:
            messagebox.showerror("Error", f"Failed to load image {i+1}.")
            return
    
    # Process the images to create the combined result
    if len(self.images) == 2:
        # Add the two images with equal weight (0.5 each)
        combined = cv2.add(
            cv2.divide(self.images[0], 2),
            cv2.divide(self.images[1], 2)
        )
        # Apply Gaussian blur
        combined = cv2.GaussianBlur(combined, (11, 11), 0)
        # Apply Otsu's thresholding
        _, combined = cv2.threshold(
            combined, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU
        )
        self.images.append(combined)
        
        # Reset layer order to default
        self.layer_order = [2, 0, 1]
        
        # Show the layer controls now that we have images
        show_layer_controls(self)
        
        # Update the display
        update_display(self)

def load_mosaic(self):
    """Load a mosaic image and either select a ROI or process full image."""
    # Verify if pixel calibration is set
    if not self.pixel_cal_input.get():
        messagebox.showwarning("Warning", "Please set the pixel calibration value.")
        return
        
    # Clear previous images
    self.proc_images = []
    self.proc_tk_images = []
    
    # Hide controls if they were previously shown
    hide_proc_layer_controls(self)
    
    # Open file dialog to select image
    file_path = fd.askopenfilename(
        title="Select a Mosaic",
        filetypes=[("Image files", "*.jpg *.jpeg *.bmp *.png *.tif *.tiff")]
    )
    
    if not file_path:
        return  # User cancelled
        
    # Load the selected image using OpenCV
    image = cv2.imread(file_path)
    if image is None:
        messagebox.showerror("Error", "Failed to load image.")
        return
        
    # Store the original image and prepare ROI image
    self.original_image = image.copy()
    self.roi_image = image

    # Hide Confirm ROI initially
    set_confirm_roi_button_visible(self, False)

    # Ask if the user wants to select a ROI
    wants_roi = messagebox.askyesno(
        "Select ROI",
        "Do you want to select a ROI (Yes) or process the entire image (No)?"
    )

    if wants_roi:
        # Enter ROI selection mode
        self.roi_mode = True

        # Bind mouse events for ROI selection
        self.proc_canvas.bind("<ButtonPress-1>", lambda event: start_roi(self, event))
        self.proc_canvas.bind("<B1-Motion>", lambda event: update_roi(self, event))
        self.proc_canvas.bind("<ButtonRelease-1>", lambda event: end_roi_drag(self, event))

        # Bind keyboard event for confirming ROI (Enter key)
        self.root.bind("<Return>", lambda event: confirm_roi(self, event))

        # Show Confirm ROI button
        set_confirm_roi_button_visible(self, True)

        # Optional guidance
        messagebox.showinfo(
            "Select ROI",
            "Click and drag to select a Region Of Interest. Press Enter or click 'Confirm ROI' when done."
        )
        
        # Now that ROI mode is enabled, update the display to show the image for ROI selection
        update_proc_display(self)
    else:
        # Process entire image as ROI
        self.roi_mode = False
        set_confirm_roi_button_visible(self, False)

        h, w = self.original_image.shape[:2]
        process_selected_roi(self, 0, 0, w, h)

    # Enable the Global Pore Stats button (allowed after load)
    self.save_mosaic_stats_data_button.config(state=tk.NORMAL)

def save_image(self):
    """Save the combined result as a TIFF file"""
    if not self.images or len(self.images) < 3:
        messagebox.showwarning("Warning", "No Binary mosaic to save.")
        return
        
    # Open file dialog to select save location
    file_path = fd.asksaveasfilename(
        title="Save Binary mosaic",
        defaultextension=".tiff",
        filetypes=[("TIFF files", "*.tiff"), ("All files", "*.*")]
    )
    
    if not file_path:
        return  # User cancelled
        
    try:
        # Save the combined result image (original resolution)
        cv2.imwrite(file_path, self.images[2])
        messagebox.showinfo("Success", f"Binary mosaic saved to '{os.path.basename(file_path)}'")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to save image: {str(e)}")

def save_proc_image(self, index):
    """Save the specified processed image as a TIFF file"""
    if not self.proc_images or index >= len(self.proc_images):
        messagebox.showwarning("Warning", f"No {self.proc_layer_names[index]} image to save.")
        return
        
    # Open file dialog to select save location
    file_path = fd.asksaveasfilename(
        title=f"Save {self.proc_layer_names[index]}",
        defaultextension=".tiff",
        filetypes=[("TIFF files", "*.tiff"), ("All files", "*.*")]
    )
    
    if not file_path:
        return  # User cancelled
        
    try:
        # Save the specified image (original resolution)
        cv2.imwrite(file_path, self.proc_images[index])
        messagebox.showinfo("Success", f"{self.proc_layer_names[index]} saved to '{os.path.basename(file_path)}'")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to save image: {str(e)}")


def save_mosaic_stats_data(self):
        # Open dialog to select save location directory
    file_path = fd.askdirectory(
        title="Select Directory to Save Global Pore Stats"
    )

    if not file_path:
        messagebox.showwarning("Warning", "No directory selected.")
        return  # User cancelled
    file_path = os.path.abspath(file_path)  # Get absolute path

    # Enter the mosaic name to append the mosaic name to the summary
    mosaic_name = sd.askstring("Mosaic Name", "Enter the name of the mosaic:")
    if not mosaic_name:
        messagebox.showwarning("Warning", "No mosaic name provided.")
        return  # User cancelled

    save_original_binary(self, file_path, mosaic_name)
    save_gpd_stats(self, file_path, mosaic_name)
    save_enhanced_contours_hdf5(self, file_path, mosaic_name)
    save_segmented_pore_data(self, file_path, mosaic_name)


def save_original_binary(self, file_path, mosaic_name):
    """Save the original binary image without contours"""
    if not hasattr(self, 'original_binary') or self.original_binary is None:
        messagebox.showwarning("Warning", "No original binary image available.")
        return
    
    # Builds the path to save the binary file
    binary_file_path = os.path.join(file_path, mosaic_name, mosaic_name + ".tiff")
    # Creates the directory if does not exist
    os.makedirs(os.path.dirname(binary_file_path), exist_ok=True) 

    try:
        # Save the original binary image
        cv2.imwrite(binary_file_path, self.original_binary)
        messagebox.showinfo("Success", f"Original binary image saved to '{os.path.basename(binary_file_path)}'")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to save image: {str(e)}")


def save_gpd_stats(self, file_path, mosaic_name):
    """Save the global pore statistics to a file"""
    if not hasattr(self, 'summary') or self.summary is None:
        messagebox.showwarning("Warning", "No global pore statistics available.")
        return
       
    try:
        # Save the summary to a Excel file using openpyxl
        if not os.path.exists(file_path + "/Global_Pore_Stats.xlsx"):
            # If file does not exist, create a new workbook and worksheet
            wb = opxl.Workbook()
            ws = wb.active
            ws.title = "Global Pore Stats"
            
            # Write headers
            headers = ["Mosaic Name",
                "Number of parent contours",
                "Number of child contours",
                "Porosity",
                "Number of parent contours <= 50μm",
                "Number of child contours\n(parent <= 50μm)",
                "Percentage of pores <= 50μm",
                "Number of parent contours > 50μm",
                "Number of child contours\n(parent > 50μm)",
                "Percentage of pores > 50μm"
            ]
            ws.append(headers)
        else:
            # If file exists, load the existing workbook and worksheet
            wb = opxl.load_workbook(file_path + "/Global_Pore_Stats.xlsx")
            ws = wb.active
        
        self.summary = (mosaic_name, ) + self.summary

        # Append the summary to the worksheet
        ws.append(self.summary)
        wb.save(file_path + "/Global_Pore_Stats.xlsx")
        wb.close()

        # Disable the Global Pore Stats button
        self.save_mosaic_stats_data_button.config(state=tk.DISABLED)
            
        messagebox.showinfo("Success", f"Global pore statistics saved to '{os.path.basename(file_path) + '/Global_Pore_Stats.xlsx'}'")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to save statistics: {str(e)}")

def save_enhanced_contours_hdf5(self, file_path, mosaic_name):
    """
    Save contours with edge information to HDF5 file.
    
    Args:
        contour_data: List of [index, is_edge, parent, [children], area, perimeter]
        filename: Output HDF5 filename
    """
    """Save the processed contours to a file"""
    if not hasattr(self, 'processed_contours') or self.processed_contours is None:
        messagebox.showwarning("Warning", "No processed contours available.")
        return
    
    # Builds the path to save the h5 file
    filename = os.path.join(file_path, mosaic_name, mosaic_name + ".h5")
    # Creates the directory if does not exist
    os.makedirs(os.path.dirname(filename), exist_ok=True) 

    contour_data = self.processed_contours

    try:
        with h5py.File(filename, 'w') as f:
            # Create a group for all contours
            contours_group = f.create_group("contours")
            
            # Store metadata about the dataset
            f.attrs['num_contours'] = len(contour_data)
                    
            # Create groups for edge and interior contours for easy filtering
            edge_group = f.create_group("edge_contours")
            interior_group = f.create_group("interior_contours")
            
            # Create a group for each contour with its index as the name for direct access
            for idx, is_edge, parent, children, area, perimeter in contour_data:
                # Use the index as the group name for direct access
                contour_group = contours_group.create_group(f"{idx}")
                
                # Store the index and edge flag as attributes
                contour_group.attrs['index'] = idx
                contour_group.attrs['is_edge'] = is_edge
                contour_group.attrs['area'] = area
                contour_group.attrs['perimeter'] = perimeter
                
                # Save parent contour
                parent_dataset = contour_group.create_dataset('parent', data=parent)
                
                # Save metadata about children
                contour_group.attrs['num_children'] = len(children)
                
                # Create a group for children
                if children:
                    children_group = contour_group.create_group('children')
                    for j, child in enumerate(children):
                        child_dataset = children_group.create_dataset(f"{j}", data=child)
                
                # Add reference to edge or interior group
                if is_edge:
                    edge_group[f"{idx}"] = contour_group.ref
                else:
                    interior_group[f"{idx}"] = contour_group.ref
        messagebox.showinfo("Success", f"Processed contours saved to '{os.path.basename(filename)}'")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to save processed contours: {str(e)}")

def save_segmented_pore_data(self, file_path, mosaic_name):
    """Save the segmented pores to xlsx fils"""
    if not hasattr(self, 'processed_cont_great_50_sz') or self.processed_cont_great_50_sz is None:
        messagebox.showwarning("Warning", "No segmented pores available.")
        return
        
    try:     
        # Segmented pores are saved according to the defined shape-size combinations.
        for shape in self.shapes:
            
            # Builds the path to save the h5 file
            filename = os.path.join(file_path, mosaic_name, mosaic_name + "_" + shape["name"] + ".xlsx")
            
            # Save the summary to a Excel file using openpyxl
            if not os.path.exists(filename):
                # If file does not exist, create a new workbook and worksheet
                wb = opxl.Workbook()
                wb.remove(wb.active)
            else:
                messagebox.showwarning("Warning", "File with segmented pore info already exist.")
                return
        
            for i, size in enumerate(self.sizes):
                # Skip invalid shape-size combinations
                if (
                    (shape["name"] != "elongated" and size["name"] in ["edS", "edM", "edL", "edXL"]) or
                    (shape["name"] != "circ" and size["name"] in ["emdS", "emdM", "emdL", "emdXL"]) or
                    ((shape["name"] not in ["circ", "MLcirc"]) and size["name"] in ["rmsS", "rmsM", "rmsL", "rmsXL"])
                ):
                    
                    ws = wb.create_sheet(size["name"], i)
                                        
                    # Write headers
                    headers = [
                        "Pore id",
                        "is_edge",
                        "Area",
                        "Perimeter",
                        "Shape",
                        "Convex Shape",
                        "Pore elongation",
                        "Irregular",
                        "Slightly irregulars",
                        "Slightly regulars",
                        "Regulars",
                        "Equivalent diameter" if size["name"] in ["edS", "edM", "edL", "edXL"] else
                        "Ellipse minor diameter" if size["name"] in ["emdS", "emdM", "emdL", "emdXL"] else
                        "Rectangle minor side",
                        None if size["name"] in ["edS", "edM", "edL", "edXL"] else
                        "Ellipse major diameter" if size["name"] in ["emdS", "emdM", "emdL", "emdXL"] else
                        "Rectangle major side",  
                        "Angle" if shape["name"] == "elongated" else None,  # Ellipse angle
                    ]
                    ws.append(headers)
                

                    try:
                        # Append the segmented pore data to the worksheet
                        for row in self.processed_cont_great_50_sz.get((shape["name"], size["name"]), []):
                            ws.append(row)
                    except Exception as e:
                        messagebox.showerror("Error", f"Failed to append data to {shape["name"]}: {str(e)}") 
                        return    

            try:
                wb.save(filename)
                wb.close()
            except Exception as e:
                messagebox.showerror("Error", f"Failed to create {filename}: {str(e)}") 
                return    
    except Exception as e:
        messagebox.showerror("Error", f"Failed to create xlsx file to save segmented pore info: {str(e)}") 
        return


                        
    messagebox.showinfo("Success", f"Segmented pore data saved to .xlsx files in'{os.path.basename(file_path) + '/' + mosaic_name}'")
                